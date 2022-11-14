import json
import random
from geopy import Nominatim
# from pymongo import MongoClient
import openpyxl
from geopy import Yandex



def get_address(type_obj):  #принимает тип объекта(соответствует названию файла), возвращает список адресов данного типа
    name_file = type_obj + ".xlsx"
    list_address = []
    workbook = openpyxl.load_workbook(name_file)
    worksheet = workbook.active
    if worksheet.max_row > 2517: #абсолютно тупая проверка, чтобы ограничить количество запросов из-за таблицы с домами
        count_row = 10  #тут ограничитель по запросам
    else:
        count_row = worksheet.max_row
    for row in worksheet.iter_rows(3, 143): #запустил на весь размер таблицы
        text = str(row[0].value)
        if text != '':
            if text.find('Address') > 0:
                list_address.append(text[text.find('Address') + len('Address'): text.find("\navailable")])
            else:
                list_address.append(text)
                #print(text)
    return list_address
#get_address('house')


def get_postamats():
    workbook = openpyxl.load_workbook("postamats.xlsx")
    worksheet = workbook.active
    data_pick = {"postamats" : []}
    i = 1
    for row in worksheet.iter_rows(2, 1644):  # запустил на весь размер таблицы
        one_pick = {"id" : i, "address" : row[4].value, "lattitude" : float(row[13].value), "longtitude" : float(row[14].value)}
        data_pick['postamats'].append(one_pick)
        i += 1
    with open("DataBase_pickpoint.json", "w") as f:
        json.dump(data_pick, f, ensure_ascii=False)
    f.close()

get_postamats()

def get_house():
    workbook = openpyxl.load_workbook("house.xlsx")
    worksheet = workbook.active
    geolocator_yandex = Yandex(api_key="1b2779ee-0b47-46e1-a95f-1f45d41033b5")
    geolocator_nomenatim = Nominatim(user_agent="memchiki")
    json_data = {}
    data_pick = {"houses": []}
    i = 1
    for row in worksheet.iter_rows(2, 302):  # запустил на весь размер таблицы
        location = geolocator_yandex.geocode(str(row[0].value))
        house = {"id": i, "address": row[4].value, "lattitude": float(row[13].value),
                    "longtitude": float(row[14].value)}
        data_pick['postamats'].append(one_pick)
        i += 1
    with open("DataBase_pickpoint.json", "w") as f:
        json.dump(data_pick, f, ensure_ascii=False)
    f.close()

class OpenMapData_PASHOK:
    def getPostamatData():
        list_object = ['mfc', 'kiosk', 'sport', 'culture', 'biblio', 'house']
        #list_object = ['house'] #только для домов
        geolocator_yandex = Yandex(api_key="1b2779ee-0b47-46e1-a95f-1f45d41033b5")
        geolocator_nomenatim = Nominatim(user_agent="memchiki")
        json_data = {}
        with open("another_safe.json", "r") as f:
            json_data = json.load(f)
        f.close()
        i = 1
        for name_object in list_object:
            list_address = get_address(name_object)
            for adress in list_address:
                try:
                    #print('adr=',adress)
                    location = geolocator_yandex.geocode(adress)
                    coef = round(random.uniform(0, 1), 4)
                    addr = geolocator_nomenatim.reverse((location.latitude, location.longitude))
                    for district in json_data["Districts"]:
                        for are in district["area"]:
                            if addr.address.find(are['name']) > 0:
                                postamat = {"id": i, "type": name_object, "lattitude": location.latitude,
                                            "longtitude": location.longitude, "address": addr.address, "coefficient": coef}
                                are['postamats'].append(postamat)
                                i += 1
                except:
                    print("Exception occured while reverse geocodein")

        with open("TmpData.json", "w") as f:
            json.dump(json_data, f, ensure_ascii=False)
        f.close()







class OpenMapData:
    def getPostamatData():
        geolocator_yandex = Yandex(api_key="1b2779ee-0b47-46e1-a95f-1f45d41033b5")
        geolocator_nomenatim = Nominatim(user_agent="memchiki")
        workbook = openpyxl.load_workbook("postamats.xlsx")
        worksheet = workbook.active
        json_data = {}
        with open("another_safe.json", "r") as f:
            json_data = json.load(f)
        f.close()
        i = 1
        for row in worksheet.iter_rows(3, 141):
            try:
                location = geolocator_yandex.geocode(row[9].value)  # номер столбца
                coef = round(random.uniform(0, 1), 4)
                addr = geolocator_nomenatim.reverse((location.latitude, location.longitude))
                for district in json_data["Districts"]:
                    for are in district["area"]:
                        if addr.address.find(are['name']) > 0:
                            postamat = {"id": i, "type": "mfc", "lattitude": location.latitude,
                                        "longtitude": location.longitude, "address": addr.address, "coefficient": coef}
                            are['postamats'].append(postamat)
                            i += 1
            except:
                print("Exception occured while reverse geocodein")

        with open("TmpData_new.json", "w") as f:
            json.dump(json_data, f, ensure_ascii=False)
        f.close()

        # client = MongoClient('localhost', 27017)
        # db = client['postamats_db']
        # db.create_collection("postamats")
        # postamats_collections = db["postamats"]
        # postamats_collections.insert_one(json_data)
        # client.close()

    def test():
        print("Try test")

        client = MongoClient('localhost', 27017)
        db = client['postamats_db']
        coll = db['postamats']
        data = coll.find_one({"Districts.name": "Северо-Западный"})
        print(data)

# OpenMapData.getPostamatData()
# OpenMapData.test()

#OpenMapData_PASHOK.getPostamatData()